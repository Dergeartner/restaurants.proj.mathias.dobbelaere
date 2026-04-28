"""
Lieferando Partner Discovery – Potsdam
======================================
Holt alle gastronomischen Betriebe in Potsdam aus OpenStreetMap (Overpass API),
berechnet einen Lead-Score, gleicht (optional) mit einer manuell erfassten
Lieferando-Partnerliste ab und schreibt eine formatierte Excel-Datei plus
JSON-Datei für das Frontend.

Lizenz der Datenquelle: ODbL (OpenStreetMap)
Bewerbungs-Mini-Projekt – kein Scraping von kommerziellen Plattformen.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
import sys
import time
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz, process

# ---------------------------------------------------------------------------
# Konfiguration
# ---------------------------------------------------------------------------

OVERPASS_URL = "https://overpass-api.de/api/interpreter"
USER_AGENT = "lieferando-potsdam-discovery/1.0 (Bewerbungsprojekt)"

# OSM-amenity-Werte, die als gastronomische Betriebe gelten
GASTRO_AMENITIES = {"restaurant", "cafe", "bar", "fast_food", "pub", "biergarten"}

# Anker-Zahl für die Story: vom Nutzer am 27.04.2026 manuell auf
# https://www.lieferando.de/lieferservice/essen/potsdam aus der Marktübersicht
# abgelesen. Keine automatische Abfrage von Lieferando.
LIEFERANDO_PARTNERS_TOTAL = 132
LIEFERANDO_SNAPSHOT_DATE = "2026-04-27"

OUTPUT_DIR = Path(__file__).parent / "output"
LIEFERANDO_LIST_FILE = Path(__file__).parent / "lieferando_partners_potsdam.txt"
LIEFERANDO_CSV_FILE = Path(__file__).parent / "lieferando_partners_potsdam.csv"
GMAPS_CSV = OUTPUT_DIR / "gmaps_potsdam.csv"

# Maximale Distanz (Meter) für GMaps-OSM-Match
GMAPS_MAX_DIST_M = 120

# Lieferando-Markenfarbe für Excel-Header
LIEFERANDO_ORANGE = "FF8000"

# ---------------------------------------------------------------------------
# Datenmodell
# ---------------------------------------------------------------------------


@dataclass
class Restaurant:
    name: str
    kategorie: str
    adresse: str
    stadtteil: str
    cuisine: str
    website: str
    telefon: str
    lat: float
    lon: float
    hat_website: bool
    lead_score: int
    auf_lieferando: bool


# ---------------------------------------------------------------------------
# Schritt 1: OSM-Daten holen
# ---------------------------------------------------------------------------


def fetch_osm_potsdam() -> list[dict[str, Any]]:
    """Holt alle gastronomischen Betriebe in Potsdam via Overpass API."""
    amenity_regex = "|".join(GASTRO_AMENITIES)
    query = f"""
    [out:json][timeout:60];
    area["name"="Potsdam"]["admin_level"="6"]->.searchArea;
    (
      node["amenity"~"^({amenity_regex})$"](area.searchArea);
      way["amenity"~"^({amenity_regex})$"](area.searchArea);
      relation["amenity"~"^({amenity_regex})$"](area.searchArea);
    );
    out center tags;
    """

    print(f"[OSM] Query an Overpass-API gesendet ...", flush=True)
    response = requests.post(
        OVERPASS_URL,
        data={"data": query},
        headers={"User-Agent": USER_AGENT},
        timeout=90,
    )
    response.raise_for_status()
    elements = response.json().get("elements", [])
    print(f"[OSM] {len(elements)} Roh-Elemente erhalten.", flush=True)
    return elements


# ---------------------------------------------------------------------------
# Schritt 2: OSM-Elemente in saubere Restaurant-Objekte umwandeln
# ---------------------------------------------------------------------------


def normalize_phone(raw: str) -> str:
    """Entfernt Leerzeichen / Bindestriche, lässt Klammern / + drin."""
    if not raw:
        return ""
    cleaned = re.sub(r"[\s\-]", "", raw.split(";")[0].strip())
    return cleaned


def build_address(tags: dict[str, str]) -> str:
    """Baut eine lesbare Adresse aus OSM-addr-Tags."""
    street = tags.get("addr:street", "").strip()
    house = tags.get("addr:housenumber", "").strip()
    plz = tags.get("addr:postcode", "").strip()
    city = tags.get("addr:city", "Potsdam").strip()

    line1 = f"{street} {house}".strip()
    line2 = f"{plz} {city}".strip()
    return ", ".join(part for part in (line1, line2) if part)


def calc_lead_score(has_website: bool, has_phone: bool) -> int:
    """3 = Website + Telefon, 2 = eines davon, 1 = nichts."""
    if has_website and has_phone:
        return 3
    if has_website or has_phone:
        return 2
    return 1


def parse_elements(elements: list[dict[str, Any]]) -> list[Restaurant]:
    """Wandelt rohe OSM-Elemente in saubere Restaurant-Objekte um."""
    restaurants: list[Restaurant] = []

    for el in elements:
        tags = el.get("tags", {})
        name = tags.get("name", "").strip()
        if not name:
            continue  # ohne Namen unbrauchbar

        # Koordinaten: bei "node" direkt lat/lon, bei "way"/"relation" der "center"
        if "lat" in el and "lon" in el:
            lat, lon = el["lat"], el["lon"]
        elif "center" in el:
            lat, lon = el["center"]["lat"], el["center"]["lon"]
        else:
            continue  # ohne Koordinaten keine Karte möglich

        kategorie = tags.get("amenity", "").strip()
        cuisine = tags.get("cuisine", "").split(";")[0].strip()
        website = tags.get("website", tags.get("contact:website", "")).strip()
        telefon = normalize_phone(tags.get("phone", tags.get("contact:phone", "")))
        stadtteil = tags.get("addr:suburb", "").strip() or "Unbekannt"

        has_website = bool(website)
        has_phone = bool(telefon)

        restaurants.append(
            Restaurant(
                name=name,
                kategorie=kategorie,
                adresse=build_address(tags),
                stadtteil=stadtteil,
                cuisine=cuisine,
                website=website,
                telefon=telefon,
                lat=lat,
                lon=lon,
                hat_website=has_website,
                lead_score=calc_lead_score(has_website, has_phone),
                auf_lieferando=False,  # wird in Schritt 3 gesetzt
            )
        )

    return restaurants


# ---------------------------------------------------------------------------
# Schritt 3: Optionaler Abgleich mit manueller Lieferando-Partnerliste
# ---------------------------------------------------------------------------


def load_lieferando_partner_names() -> list[str]:
    """Liest manuell erfasste Lieferando-Partnernamen (eine pro Zeile, .txt)."""
    if not LIEFERANDO_LIST_FILE.exists():
        return []
    with LIEFERANDO_LIST_FILE.open("r", encoding="utf-8") as f:
        return [line.strip() for line in f if line.strip() and not line.startswith("#")]


def load_lieferando_partner_csv() -> list[dict]:
    """Liest manuell erfasste Lieferando-Partner mit Name + Adresse aus CSV.

    Erwartete Spalten: name, adresse (ggf. plz, strasse separat — flexibel).
    Returns: [{name, adresse, plz, strasse}, ...]
    """
    import csv as _csv
    if not LIEFERANDO_CSV_FILE.exists():
        return []
    rows: list[dict] = []
    with LIEFERANDO_CSV_FILE.open("r", encoding="utf-8", newline="") as f:
        # Auto-detect Delimiter
        sample = f.read(2048)
        f.seek(0)
        try:
            dialect = _csv.Sniffer().sniff(sample, delimiters=",;|\t")
        except _csv.Error:
            dialect = _csv.excel
        reader = _csv.DictReader(f, dialect=dialect)
        for row in reader:
            normalized = {k.lower().strip(): (v or "").strip() for k, v in row.items() if k}
            name = normalized.get("name", "")
            if not name:
                continue
            adresse = normalized.get("adresse") or normalized.get("address") or ""
            plz = normalized.get("plz") or normalized.get("postcode") or ""
            strasse = normalized.get("strasse") or normalized.get("street") or ""
            # Falls nur "adresse" gefüllt, PLZ + Straße daraus parsen
            if adresse and not plz:
                m = re.search(r"\b(\d{5})\b", adresse)
                if m:
                    plz = m.group(1)
            if adresse and not strasse:
                # Alles vor der PLZ ist die Straße + Hausnr
                strasse_raw = re.split(r"\d{5}", adresse)[0].strip().rstrip(",").strip()
                strasse = strasse_raw
            rows.append({
                "name": name,
                "adresse": adresse,
                "plz": plz,
                "strasse": strasse,
            })
    return rows


def _strassenname_only(strasse: str) -> str:
    """Aus 'Mittelstraße 20' wird 'mittelstrasse'. Hausnummer raus, lowercase, ß→ss."""
    s = strasse.lower().replace("ß", "ss")
    s = re.sub(r"\d+\s*[a-z]?", " ", s)  # Hausnummern entfernen
    s = re.sub(r"[^a-zäöü ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def clean_restaurant_name(name: str) -> str:
    """Säubert Restaurant-Namen für robusteren Fuzzy-Match.

    Entfernt:
    - Trademark-Symbole (®, ™, ©)
    - Stadt/Stadtteil-Suffixe wie "| Potsdam", "I Potsdam-Babelsberg", "- Potsdam-West"
    - Trailing Stadtteil-Tags wie "-Babelsberg", "-West", "Alt Nowawes"
    - Whitespace + Trailing-Sonderzeichen
    """
    if not name:
        return ""
    s = name
    # Trademark-Symbole + Sonder-Whitespaces
    s = s.replace("®", "").replace("™", "").replace("©", "")
    s = s.replace("­", "")  # weiches Trennzeichen (z.B. "Pots­dam")

    # Punkt-Pipe-I-Slash-Trenner zu "Stadt-Tag"
    SEP = r"[\s\-\|/]+"
    POTSDAM_TAGS = (
        r"Potsdam(?:\s*[-\s]\s*(?:Hbf|Hauptbahnhof|Mitte|West|Ost|Nord|S(?:ü|u)d|"
        r"Babelsberg|Alt[\s\-]+Nowawes|Innenstadt|Bornstedt))?"
    )
    # "X | Potsdam-Babelsberg" → "X"
    s = re.sub(rf"\s*[\|]\s*{POTSDAM_TAGS}\s*$", "", s, flags=re.IGNORECASE)
    s = re.sub(rf"\s+I\s+{POTSDAM_TAGS}\s*$", "", s, flags=re.IGNORECASE)
    s = re.sub(rf"\s*-\s*{POTSDAM_TAGS}\s*$", "", s, flags=re.IGNORECASE)
    s = re.sub(rf"\s+{POTSDAM_TAGS}\s*$", "", s, flags=re.IGNORECASE)

    # Auch ohne "Potsdam"-Wort: nackte Stadtteil-Suffixe entfernen
    s = re.sub(
        r"\s*-\s*(Babelsberg|West|Ost|Nord|S(?:ü|u)d|Innenstadt|Bornstedt|Drewitz)\s*$",
        "",
        s,
        flags=re.IGNORECASE,
    )

    # & in "und" normalisieren — hilft bei "Sushi & Bowls" vs "Sushi und Bowls"
    s = re.sub(r"\s*&\s*", " und ", s)

    # Doppel-Whitespace + Trailing Sonderzeichen
    s = re.sub(r"\s+", " ", s).strip().rstrip("|").rstrip("-").strip()
    return s


# Stadtteile/Bezirke außerhalb Potsdams, die aus der Lieferando-Liste rausgefiltert
# werden sollen (ihre Lieferzone reicht zwar bis Potsdam, sie sind aber nicht in Potsdam)
NON_POTSDAM_LOCATION_KEYWORDS = [
    "berlin", "steglitz", "wedding", "kreuzberg", "neukölln", "neukoelln",
    "michendorf",  # Michendorf ist eigene Gemeinde, gehört nicht zu Potsdam
]


def is_non_potsdam_partner(name: str) -> bool:
    """Heuristik: Restaurant gehört vermutlich nicht zu Potsdam, sondern zu einer
    Nachbargemeinde / Berlin / etc. — sollte aus der Marktanalyse rausfallen."""
    n = name.lower()
    return any(kw in n for kw in NON_POTSDAM_LOCATION_KEYWORDS)


# Lieferando-Listings, die zwar Partner sind, aber keine OSM-Gastrobetriebe sein können
# (Apotheken, Supermärkte, Spätkäufe) – die NICHT in OSM matchen sollen, aber wir
# dokumentieren sie trotzdem als auf_lieferando=true zur vollständigen Markt-Repräsentation.
NON_GASTRO_LIEFERANDO_KEYWORDS = [
    "apotheke", "rewe", "flink", "iqos", "kiosk", "spätkauf", "spaetkauf",
    "weinfreunde", "alieman", "lebensmittel", "manufaktur",
]


def is_non_gastro_partner(name: str) -> bool:
    n = name.lower()
    return any(kw in n for kw in NON_GASTRO_LIEFERANDO_KEYWORDS)


def match_lieferando_csv(
    restaurants: list[Restaurant], partners: list[dict]
) -> tuple[int, list[dict]]:
    """Match Lieferando-Partner gegen OSM-Liste mit Name-Cleaning + optional Adresse.

    Returns: (anzahl_matches, liste_unmatched_partners)
    """
    if not partners:
        return 0, []

    # Cleaning-Lookup: clean_name → original name
    osm_clean_to_name: dict[str, str] = {}
    for r in restaurants:
        cleaned = clean_restaurant_name(r.name).lower()
        if cleaned and cleaned not in osm_clean_to_name:
            osm_clean_to_name[cleaned] = r.name
    osm_by_name = {r.name: r for r in restaurants}
    osm_clean_keys = list(osm_clean_to_name.keys())

    matched = 0
    unmatched: list[dict] = []

    for partner in partners:
        partner_clean = clean_restaurant_name(partner["name"]).lower()
        if not partner_clean:
            unmatched.append(partner)
            continue

        # Direct-Match (sehr häufig nach Cleaning)
        if partner_clean in osm_clean_to_name:
            target = osm_by_name[osm_clean_to_name[partner_clean]]
            if not target.auf_lieferando:
                target.auf_lieferando = True
                matched += 1
            continue

        # Top-5 Fuzzy-Kandidaten gegen geputzte Namen
        candidates = process.extract(
            partner_clean, osm_clean_keys, scorer=fuzz.token_set_ratio, limit=5
        )
        partner_strasse_norm = _strassenname_only(partner.get("strasse", ""))
        partner_plz = partner.get("plz", "")

        best = None
        for cand_clean, name_score, _ in candidates:
            if name_score < 75:
                continue
            r = osm_by_name[osm_clean_to_name[cand_clean]]
            r_strasse_norm = _strassenname_only(
                r.adresse.split(",")[0] if r.adresse else ""
            )
            r_plz = ""
            m = re.search(r"\b(\d{5})\b", r.adresse or "")
            if m:
                r_plz = m.group(1)

            # PLZ-Match (stark) — Schwelle 75 reicht
            if partner_plz and r_plz and partner_plz == r_plz and name_score >= 75:
                best = r
                break
            # Straßen-Fuzzy
            if partner_strasse_norm and r_strasse_norm:
                strasse_score = fuzz.token_set_ratio(
                    partner_strasse_norm, r_strasse_norm
                )
                if strasse_score >= 70 and name_score >= 78:
                    best = r
                    break
            # Fallback: nur Name, höhere Schwelle
            if name_score >= 88:
                best = r
                break
            # Sehr-hohe-Name-Schwelle für Stadt-Suffix-Cases
            if name_score >= 82 and len(partner_clean) >= 8:
                best = r
                break

        if best:
            if not best.auf_lieferando:
                best.auf_lieferando = True
                matched += 1
        else:
            unmatched.append(partner)

    return matched, unmatched


def add_unmatched_partners_as_restaurants(
    restaurants: list[Restaurant], unmatched: list[dict]
) -> int:
    """Fügt Lieferando-Partner, die nicht in OSM gefunden wurden, als neue Einträge
    in die Liste ein. Damit hat jeder Partner mindestens eine Repräsentation.

    Lat/Lon werden auf Potsdam-Center gesetzt (52.4009, 13.0591) damit die Karte
    sie clustert. Sie haben kategorie='lieferando_only' und kein OSM-Pendant.
    """
    if not unmatched:
        return 0

    # Doppelte vermeiden — gleichen Namen nicht 2× hinzufügen
    existing_clean = {clean_restaurant_name(r.name).lower() for r in restaurants}

    POTSDAM_CENTER_LAT = 52.4009
    POTSDAM_CENTER_LON = 13.0591

    added = 0
    skipped_non_potsdam = 0
    for p in unmatched:
        cleaned = clean_restaurant_name(p["name"]).lower()
        if cleaned in existing_clean:
            continue
        # Restaurants außerhalb Potsdams (Berlin/Steglitz/Michendorf etc.) raus
        if is_non_potsdam_partner(p["name"]):
            skipped_non_potsdam += 1
            continue
        existing_clean.add(cleaned)

        # Kategorie: nicht-Gastro (Apotheke, Supermarkt etc.) eigenständig markieren
        kategorie = "lieferando_non_gastro" if is_non_gastro_partner(p["name"]) else "lieferando_only"

        # Slight jitter um Centroid damit Pins auf Karte nicht überlappen
        import random
        jitter_lat = (random.random() - 0.5) * 0.02
        jitter_lon = (random.random() - 0.5) * 0.02

        new_r = Restaurant(
            name=p["name"],
            kategorie=kategorie,
            adresse=p.get("adresse", ""),
            stadtteil="Lieferando-only",
            cuisine="",
            website="",
            telefon="",
            lat=POTSDAM_CENTER_LAT + jitter_lat,
            lon=POTSDAM_CENTER_LON + jitter_lon,
            hat_website=False,
            lead_score=1,
            auf_lieferando=True,
        )
        restaurants.append(new_r)
        added += 1

    if skipped_non_potsdam:
        print(
            f"[Skip] {skipped_non_potsdam} Lieferando-Listings außerhalb Potsdams "
            f"(Berlin/Michendorf/etc.) ignoriert.",
            flush=True,
        )
    return added


# ---------------------------------------------------------------------------
# Schritt 2b: Fehlende Stadtteile per Geo-Lookup auffüllen
# ---------------------------------------------------------------------------


def fetch_potsdam_suburbs() -> list[tuple[str, float, float]]:
    """Holt alle Potsdam-Stadtteile als Nodes mit Centroid-Koordinaten.
    Returns: Liste von (name, lat, lon) — Stadtteile sind in Potsdam meist als
    place=suburb/quarter/neighbourhood Nodes getaggt, nicht als Polygone.
    """
    query = """
    [out:json][timeout:60];
    area["name"="Potsdam"]["admin_level"="6"]->.a;
    (
      node["place"="suburb"](area.a);
      node["place"="quarter"](area.a);
      node["place"="neighbourhood"](area.a);
    );
    out tags center;
    """
    print("[Stadtteil] Hole Stadtteil-Centroide von Overpass ...", flush=True)
    response = requests.post(
        OVERPASS_URL,
        data={"data": query},
        headers={"User-Agent": USER_AGENT},
        timeout=90,
    )
    response.raise_for_status()
    data = response.json()

    suburbs: list[tuple[str, float, float]] = []
    for el in data.get("elements", []):
        name = el.get("tags", {}).get("name", "").strip()
        if not name:
            continue
        if "lat" in el and "lon" in el:
            suburbs.append((name, el["lat"], el["lon"]))
        elif "center" in el:
            suburbs.append((name, el["center"]["lat"], el["center"]["lon"]))
    print(f"[Stadtteil] {len(suburbs)} Stadtteil-Centroide geladen.", flush=True)
    return suburbs


def assign_stadtteile(
    restaurants: list[Restaurant],
    suburbs: list[tuple[str, float, float]],
) -> int:
    """Setzt für alle Restaurants mit unbekanntem Stadtteil den nächstgelegenen
    Stadtteil (Nearest-Neighbor über Haversine-Distanz). Returns: Anzahl gefüllt.
    """
    if not suburbs:
        return 0

    filled = 0
    for r in restaurants:
        if r.stadtteil and r.stadtteil != "Unbekannt":
            continue
        best_name = "Unbekannt"
        best_dist = float("inf")
        for name, slat, slon in suburbs:
            # Approximation reicht völlig für Nearest-Neighbor: keine Haversine nötig
            d = (r.lat - slat) ** 2 + (r.lon - slon) ** 2
            if d < best_dist:
                best_dist = d
                best_name = name
        if best_name != "Unbekannt":
            r.stadtteil = best_name
            filled += 1
    print(f"[Stadtteil] {filled} unbekannte Stadtteile aufgefüllt.", flush=True)
    return filled


# ---------------------------------------------------------------------------
# Optionaler Schritt 3b: GMaps-CSV anreichern (Telefon, Website)
# ---------------------------------------------------------------------------


def haversine_m(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Großkreis-Distanz zwischen zwei Punkten in Metern."""
    R = 6_371_000
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * R * math.asin(math.sqrt(a))


def normalize_for_match(s: str) -> str:
    """Lowercase + Sonderzeichen raus für besseres Fuzzy-Matching."""
    s = s.lower()
    s = re.sub(r"[^a-zäöüß0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def enrich_with_gmaps(restaurants: list[Restaurant], csv_path: Path) -> tuple[int, int]:
    """Mergt GMaps-CSV in die OSM-Liste. Match-Strategie:
    1. Geo-Distanz < 120m UND Name-Token-Set-Ratio >= 70  -> Match
    2. Geo-Distanz < 50m  (auch ohne Name-Match)          -> Match (gleicher Ort)

    Befüllt fehlende telefon/website-Felder. Verändert keine vorhandenen Werte.

    Returns: (anzahl_matches, anzahl_telefon_neu)
    """
    if not csv_path.exists():
        print(f"[GMaps] {csv_path.name} nicht gefunden — überspringe Anreicherung.")
        return 0, 0

    with open(csv_path, "r", encoding="utf-8", errors="replace", newline="") as f:
        gmaps_rows = list(csv.DictReader(f))

    if not gmaps_rows:
        print("[GMaps] CSV leer.")
        return 0, 0

    matches = 0
    new_phones = 0
    new_websites = 0

    for g in gmaps_rows:
        title = (g.get("title") or "").strip()
        if not title:
            continue
        try:
            glat = float(g.get("latitude") or 0)
            glon = float(g.get("longitude") or 0)
        except ValueError:
            continue
        if glat == 0 or glon == 0:
            continue

        gnorm = normalize_for_match(title)
        best_idx = -1
        best_dist = float("inf")
        best_score = 0

        for i, r in enumerate(restaurants):
            dist = haversine_m(glat, glon, r.lat, r.lon)
            if dist > GMAPS_MAX_DIST_M:
                continue
            score = fuzz.token_set_ratio(gnorm, normalize_for_match(r.name))
            if dist < 50 or score >= 70:
                if dist < best_dist:
                    best_dist = dist
                    best_score = score
                    best_idx = i

        if best_idx < 0:
            continue

        target = restaurants[best_idx]
        gphone = (g.get("phone") or "").strip()
        gweb = (g.get("website") or "").strip()
        changed = False

        if gphone and not target.telefon:
            target.telefon = normalize_phone(gphone)
            new_phones += 1
            changed = True
        if gweb and not target.website:
            target.website = gweb
            new_websites += 1
            changed = True

        if changed:
            target.hat_website = bool(target.website)
            target.lead_score = calc_lead_score(target.hat_website, bool(target.telefon))
            matches += 1

    print(
        f"[GMaps] {matches} Anreicherungen "
        f"(+{new_phones} Telefonnummern, +{new_websites} Websites)."
    )
    return matches, new_phones


# ---------------------------------------------------------------------------
# Schritt 4: Outputs schreiben (JSON für Frontend, Excel für Sales)
# ---------------------------------------------------------------------------


def write_json(restaurants: list[Restaurant]) -> Path:
    """Schreibt schlankes JSON für das Frontend."""
    payload = {
        "generated_at": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "lieferando_partners_total": LIEFERANDO_PARTNERS_TOTAL,
        "lieferando_snapshot_date": LIEFERANDO_SNAPSHOT_DATE,
        "restaurants": [asdict(r) for r in restaurants],
    }
    out_path = OUTPUT_DIR / "restaurants.json"
    out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return out_path


def style_header(ws, num_columns: int) -> None:
    """Header in Lieferando-Orange, weiße fette Schrift, Zeile freezen."""
    fill = PatternFill("solid", fgColor=LIEFERANDO_ORANGE)
    font = Font(bold=True, color="FFFFFF")
    for col in range(1, num_columns + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def autosize(ws) -> None:
    """Naive Spaltenbreite basierend auf längstem Wert."""
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 50)


def add_lead_score_formatting(ws, score_col_letter: str, last_row: int) -> None:
    """Ampel-Färbung für Lead_Score-Spalte: rot (1) / gelb (2) / grün (3)."""
    rng = f"{score_col_letter}2:{score_col_letter}{last_row}"
    red = PatternFill("solid", fgColor="F8CBAD")
    yellow = PatternFill("solid", fgColor="FFE699")
    green = PatternFill("solid", fgColor="C6EFCE")
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=["1"], fill=red))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=["2"], fill=yellow))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=["3"], fill=green))


def write_excel(restaurants: list[Restaurant]) -> Path:
    """Schreibt formatierte Excel-Datei mit 4 Sheets."""
    wb = Workbook()

    # ---- Sheet 1: Partnerliste -------------------------------------------
    ws1 = wb.active
    ws1.title = "Partnerliste"
    headers = [
        "Name",
        "Kategorie",
        "Adresse",
        "Stadtteil",
        "Cuisine",
        "Website",
        "Telefon",
        "Hat_Website",
        "Lead_Score",
        "Auf_Lieferando",
        "Notizen",
    ]
    ws1.append(headers)
    for r in restaurants:
        ws1.append([
            r.name,
            r.kategorie,
            r.adresse,
            r.stadtteil,
            r.cuisine,
            r.website,
            r.telefon,
            r.hat_website,
            r.lead_score,
            r.auf_lieferando,
            "",
        ])
    style_header(ws1, len(headers))
    add_lead_score_formatting(ws1, "I", ws1.max_row)
    ws1.auto_filter.ref = ws1.dimensions
    autosize(ws1)

    # ---- Sheet 2: Marktanalyse -------------------------------------------
    ws2 = wb.create_sheet("Marktanalyse")
    total = len(restaurants)
    with_web = sum(1 for r in restaurants if r.hat_website)
    hot_leads = sum(1 for r in restaurants if r.lead_score == 3)
    on_lieferando = sum(1 for r in restaurants if r.auf_lieferando)

    ws2.append([f"Gastronomische Betriebe in Potsdam: {total}"])
    ws2.append([f"davon mit Website: {with_web} ({_pct(with_web, total)})"])
    ws2.append([f"Hot Leads (Score = 3): {hot_leads}"])
    ws2.append([f"Lieferando-Partner gesamt (Marktübersicht {LIEFERANDO_SNAPSHOT_DATE}): {LIEFERANDO_PARTNERS_TOTAL}"])
    ws2.append([f"davon in OSM-Liste gemappt: {on_lieferando}"])
    ws2.append([f"Marktdurchdringung Lieferando: ~{_pct(LIEFERANDO_PARTNERS_TOTAL, total)}"])
    ws2.append([])

    _write_pivot(ws2, "Anzahl pro Stadtteil", _count_by(restaurants, lambda r: r.stadtteil))
    _write_pivot(ws2, "Anzahl pro Kategorie", _count_by(restaurants, lambda r: r.kategorie))
    _write_pivot(
        ws2,
        "Top 15 Cuisines",
        _count_by(restaurants, lambda r: r.cuisine or "(nicht gesetzt)"),
        limit=15,
    )
    _write_pivot(
        ws2,
        "Website-Coverage",
        [("mit Website", with_web), ("ohne Website", total - with_web)],
    )
    autosize(ws2)

    # ---- Sheet 3: Hot Leads ----------------------------------------------
    ws3 = wb.create_sheet("Hot Leads")
    ws3.append(headers)
    hot = sorted(
        [r for r in restaurants if r.lead_score >= 2],
        key=lambda r: (-r.lead_score, r.name.lower()),
    )
    for r in hot:
        ws3.append([
            r.name, r.kategorie, r.adresse, r.stadtteil, r.cuisine, r.website,
            r.telefon, r.hat_website, r.lead_score, r.auf_lieferando, "",
        ])
    style_header(ws3, len(headers))
    if ws3.max_row > 1:
        add_lead_score_formatting(ws3, "I", ws3.max_row)
        ws3.auto_filter.ref = ws3.dimensions
    autosize(ws3)

    # ---- Sheet 4: Akquise-Pipeline (Nicht-Partner mit Lead-Score >= 2) ----
    ws4 = wb.create_sheet("Akquise-Pipeline")
    ws4.append(headers)
    akquise = sorted(
        [r for r in restaurants if not r.auf_lieferando and r.lead_score >= 2],
        key=lambda r: (-r.lead_score, r.name.lower()),
    )
    for r in akquise:
        ws4.append([
            r.name, r.kategorie, r.adresse, r.stadtteil, r.cuisine, r.website,
            r.telefon, r.hat_website, r.lead_score, r.auf_lieferando, "",
        ])
    style_header(ws4, len(headers))
    if ws4.max_row > 1:
        add_lead_score_formatting(ws4, "I", ws4.max_row)
        ws4.auto_filter.ref = ws4.dimensions
    autosize(ws4)

    out_path = OUTPUT_DIR / "partnerliste_potsdam.xlsx"
    wb.save(out_path)
    return out_path


def _count_by(restaurants: list[Restaurant], key) -> list[tuple[str, int]]:
    counts: dict[str, int] = {}
    for r in restaurants:
        k = key(r) or "(nicht gesetzt)"
        counts[k] = counts.get(k, 0) + 1
    return sorted(counts.items(), key=lambda kv: -kv[1])


def _write_pivot(ws, title: str, rows: list[tuple[str, int]], limit: int | None = None) -> None:
    ws.append([title])
    title_cell = ws.cell(row=ws.max_row, column=1)
    title_cell.font = Font(bold=True)
    for label, count in rows[: limit or len(rows)]:
        ws.append([label, count])
    ws.append([])


def _pct(part: int, total: int) -> str:
    return f"{(100 * part / total):.1f} %" if total else "0 %"


# ---------------------------------------------------------------------------
# Hauptablauf
# ---------------------------------------------------------------------------


def main() -> int:
    parser = argparse.ArgumentParser(description="OSM-Scrape + optionales GMaps-Merge")
    parser.add_argument(
        "--merge-gmaps",
        action="store_true",
        help="Mergt output/gmaps_potsdam.csv in die OSM-Liste (vorher enrich_gmaps.py laufen lassen).",
    )
    args = parser.parse_args()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    elements = fetch_osm_potsdam()
    restaurants = parse_elements(elements)
    print(f"[Parser] {len(restaurants)} Betriebe mit Namen + Koordinaten erfasst.", flush=True)

    # Fehlende Stadtteile via Polygon-Lookup auffüllen (immer ausführen)
    try:
        suburbs = fetch_potsdam_suburbs()
        assign_stadtteile(restaurants, suburbs)
    except Exception as e:
        print(f"[Stadtteil] Fehler ({e}) – überspringe Auffüllung.", flush=True)

    if args.merge_gmaps:
        enrich_with_gmaps(restaurants, GMAPS_CSV)

    # CSV (mit Adresse) bevorzugen — robuster als nur Namen
    partners_csv = load_lieferando_partner_csv()
    if not partners_csv:
        partner_names = load_lieferando_partner_names()
        partners_csv = [{"name": n, "adresse": "", "plz": "", "strasse": ""} for n in partner_names]

    if partners_csv:
        matched, unmatched = match_lieferando_csv(restaurants, partners_csv)
        added = add_unmatched_partners_as_restaurants(restaurants, unmatched)
        print(
            f"[Match] {matched}/{len(partners_csv)} Lieferando-Partner in OSM-Liste gefunden, "
            f"{added} als Lieferando-only-Einträge ergänzt.",
            flush=True,
        )
    else:
        print(
            f"[Match] Keine Lieferando-Partnerliste gefunden ({LIEFERANDO_CSV_FILE.name} oder {LIEFERANDO_LIST_FILE.name}).",
            flush=True,
        )

    json_path = write_json(restaurants)
    print(f"[Output] JSON: {json_path}", flush=True)
    excel_path = write_excel(restaurants)
    print(f"[Output] Excel: {excel_path}", flush=True)

    # Kurz-Insights für die Konsole / Bewerbung
    total = len(restaurants)
    with_web = sum(1 for r in restaurants if r.hat_website)
    hot = sum(1 for r in restaurants if r.lead_score == 3)
    not_partner = sum(1 for r in restaurants if not r.auf_lieferando)
    print()
    print("=" * 60)
    print(f"Potsdam-Markt-Snapshot")
    print("=" * 60)
    print(f"Gastronomische Betriebe (OSM):     {total}")
    print(f"davon mit Website:                 {with_web} ({_pct(with_web, total)})")
    print(f"Hot Leads (Website + Telefon):     {hot}")
    print(f"Lieferando-Partner (Snapshot):     {LIEFERANDO_PARTNERS_TOTAL}")
    print(f"Marktdurchdringung Lieferando:     ~{_pct(LIEFERANDO_PARTNERS_TOTAL, total)}")
    print(f"Akquise-Universe (Nicht-Partner):  {not_partner}")
    print("=" * 60)

    return 0


if __name__ == "__main__":
    sys.exit(main())
